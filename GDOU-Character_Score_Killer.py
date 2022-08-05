#-*-coding:utf-8-*-
import easygui as g
import os
import openpyxl
import string
import webbrowser
import shutil
import datetime

#
version = "1.2.1"
swName = "广东海洋大学品行分统计杀手_v" + version
#
class Excel:
    def __init__(self,filePath,sheet):
        self.ExcelOpen = openpyxl.load_workbook(filePath)
        self.SheetOpen = self.ExcelOpen[sheet]
    def GetData(self,min_row, max_row, min_col, max_col):
        rawLoop = 0
        sheetData = {}
        for row in self.SheetOpen.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            rowlist = {}
            i = 0
            for cell in row:
                rowlist[i] = cell.value
                i += 1
            sheetData[rawLoop] = rowlist
            rawLoop += 1
        return sheetData
    def GetDataFilter(self,min_row,max_row,min_col,max_col,conditionKey,conditionValue):
        rawLoop = 0
        sheetData = {}
        for row in self.SheetOpen.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            rowlist = {}
            i = 0
            for cell in row:
                rowlist[i] = cell.value
                i += 1
            if (rowlist[conditionKey] == conditionValue):
                sheetData[rawLoop] = rowlist
                rawLoop += 1
        return sheetData
    def GetRowMax(self):
        return self.SheetOpen.max_row
    def GetColMax(self):
        return  self.SheetOpen.max_column
    def InsertValue(self,row,col,v):
        self.SheetOpen.cell(row,col).value = v
    def InsertNum(self,row,col,v):
        self.SheetOpen.cell(row, col).value = v
        self.SheetOpen.cell(row,col).number_format = '0'
    def Save(self,filePath):
        self.ExcelOpen.save(filePath)

class File:
    def __init__(self,filePath):
        self.FilePath = filePath
    def Cheak(self):
        return os.path.exists(self.FilePath)
    def Write(self,s):
        if (os.path.exists(os.path.dirname(self.FilePath)) == False):
            os.mkdir(os.path.dirname(self.FilePath))
        fileOpen = open(self.FilePath,"w")
        fileOpen.write(s)
        fileOpen.close()
    def Read(self):
        self.FileOpen = open(self.FilePath,"r")
        return self.FileOpen.read()
    def ReadClose(self):
        self.FileOpen.close()
    def ReadAC(self):
        self.FileOpen = open(self.FilePath, "r")
        fileContent = self.FileOpen.read()
        self.FileOpen.close()
        return fileContent

def ImportDataBasic():
    g.msgbox("接下来请导入含有你班级同学基本信息的Excel表格\n该Excel表格严格要求如图的格式依次写入班级同学们的基本信息", swName, "好哒", "../img/help1.gif")
    while True:
        while True:
            dataBasic_file = g.fileopenbox("导入含有班级学生基本信息的Excel表格", swName)
            if (dataBasic_file == None):
                return False
            else:
                try:
                    '''
                    中歌院播音1212班长“顾客点了一份炒蛋，酒吧炸了”这句话立马激起我把判断用户导入是否是Excel表格通过后缀“.xlsx”来识别的伪判断改为读取文件的真判断
                    if (dataBasic_file.endswith(".xlsx") == False):
                        ...
                    '''
                    excelBasic = Excel(dataBasic_file, "Sheet1")
                except:
                    g.msgbox("请导入Excel表格！", swName, "重新导入")
                else:
                    break
        if (excelBasic.GetRowMax() == 1):
            g.msgbox("导入的Excel表格未含有任何学生信息或未严格按要求格式写入学生信息！", swName, "重新导入")
        else:
            shutil.copyfile(dataBasic_file, "../config/classBasicData.xlsx")
            return True

def ImportClassName():
    while True:
        className = g.enterbox("请输入你的班级",swName,"(例如：能源1214)")
        if (className == None):
            g.msgbox("请输入你的班级！",swName,"好吧")
        else:
            className_file = File("../config/className.properties")
            className_file.Write(className)
            break

def SetShowDataInclude():
    while True:
        showDataInclude = g.choicebox("最终生成表格统计数据包含的学生范围",swName,("包含全部同学（不管Ta是否有品行分）","仅包含有品行分的同学"))
        if(showDataInclude == None):
            g.msgbox("请设定一个范围！",swName,"好吧")
        elif(showDataInclude == '包含全部同学（不管Ta是否有品行分）'):
            showDataInclude_file = File("../config/showDataInclude.properties")
            showDataInclude_file.Write("0")
            break
        elif(showDataInclude == '仅包含有品行分的同学'):
            showDataInclude_file = File("../config/showDataInclude.properties")
            showDataInclude_file.Write("1")
            break

def EventFormat(id, event, score):
    global space
    count_en = count_dg = count_zh = count_puF = count_puH = 0
    for word in event:
        if word in string.ascii_letters:
            count_en += 1
        elif word.isdigit():
            count_dg += 1
        elif word.isalpha():
            count_zh += 1
        else:
            if (word == "”" or word =="‘"):
                count_puF += 1
            else:
                count_puH += 1
    eventStrCount = (count_en + count_dg + count_puH + ((count_zh + count_puF) * 2))
    num_space = 36 - eventStrCount
    while (num_space < 0):
        num_space += 36
    i = 0
    while (i < num_space):
        if (i == 0):
            space = ' '
        else:
            space = space + ' '
        i += 1
    return str(id) + '.' + event + space + str(score)

def Ac():
    while True:
        g.msgbox("微信扫码关注“奥德赛创始ODC”公众号及时收到我们充满创意产品的最新消息\n也许我们未来还会制作像本软件一样解放双手的产品\n本软件版本更新也会在公众号同步发布\n\n关注后发送信息“广东海洋大学品行分统计杀手“给公众号即可免费获取激活码",swName,"已获取激活码开始输入激活码","../img/odcori.gif")
        acword = g.enterbox("请输入激活码",swName)
        if (acword == None):
            break
            print(acword)
        elif(acword != "ODCGDOUCSK"):
            g.msgbox("激活码错误！",swName,"重新输入")
        else:
            ac_file = File("C://Users//Public//Documents//system_//DDS9QHcXbpdMjNzyOxRI1sKv3ugLmqe2JVnZ.properties")
            ac_file.Write("l2Ah5PEqxdNbaZtmeJWDi86FGCjnkYu1UIQTLsRv\n9xQYcVe56Dz0vGsLS3jWkryMIN4fEamn8X1pPhOA\nbfynRCNDuhBvXV6e7LP52pat0j134dOlKc9zoUZH\nWDSRL2G8XUripwM5VnE60zAq1Ihf7eBvl9JaZkso\n1xNiMBpQ7rjqP5wb9DLVIFoYy4su6cKJdUmOA2ET\ngepuh5i0aonMsUfJXwZz4A69NqOQTPyHGCr7ISdL\n9svLRldfiyXQrtHWTVZ078k6Ppb5FmgU14KMNDcC\nTpizKn5dAvrQFDYbymIGec8UaMtlZ6Ej10Pu3RSL\nQYGIvMqEjKg3PVt2s0p9BCT8Rx764ZnewXDzJmUd\naR9oxVrTJIU8mGls7kcPuDXF41hvEYzAW36HKeLB")
            break
def Ac_ban():
    choice = g.ccbox("试用次数用完啦！", swName, ("免费激活软件", "关闭软件"))
    if (choice == False or choice == None):
        os._exit(0)
    else:
        while True:
            Ac()
            ac_file = File("C://Users//Public//Documents//system_//DDS9QHcXbpdMjNzyOxRI1sKv3ugLmqe2JVnZ.properties")
            if (ac_file.ReadAC() == "l2Ah5PEqxdNbaZtmeJWDi86FGCjnkYu1UIQTLsRv\n9xQYcVe56Dz0vGsLS3jWkryMIN4fEamn8X1pPhOA\nbfynRCNDuhBvXV6e7LP52pat0j134dOlKc9zoUZH\nWDSRL2G8XUripwM5VnE60zAq1Ihf7eBvl9JaZkso\n1xNiMBpQ7rjqP5wb9DLVIFoYy4su6cKJdUmOA2ET\ngepuh5i0aonMsUfJXwZz4A69NqOQTPyHGCr7ISdL\n9svLRldfiyXQrtHWTVZ078k6Ppb5FmgU14KMNDcC\nTpizKn5dAvrQFDYbymIGec8UaMtlZ6Ej10Pu3RSL\nQYGIvMqEjKg3PVt2s0p9BCT8Rx764ZnewXDzJmUd\naR9oxVrTJIU8mGls7kcPuDXF41hvEYzAW36HKeLB"):
                g.msgbox("激活成功！您现在可以永久使用",swName,"好哒")
                break
            else:
                choice = g.ccbox("激活失败！",swName,("重新激活","关闭软件"))
                if (choice == False or choice == None):
                    os._exit(0)
def Ad():
    webbrowser.open("https://github.com/leoweyr/GDOU-Character_Score_Killer")
    g.msgbox("如果你也是开发者的话，就给这个项目一个Star吧", swName, "已Star")
    webbrowser.open("https://space.bilibili.com/381580563")
    g.msgbox("如果本软件对你很有用的话一定要关注李燊同学的B站哟，这是对他创作的最大支持",swName,"已对他任意视频一键三连")
    g.msgbox("微信扫码关注“奥德赛创始ODC”公众号及时收到我们充满创意产品的最新消息\n也许我们未来还会制作像本软件一样解放双手的产品\n本软件版本更新也会在公众号同步发布",swName,"已微信扫码并关注公众号","../img/odcori.gif")
#
ac_file = File("C://Users//Public//Documents//system_//DDS9QHcXbpdMjNzyOxRI1sKv3ugLmqe2JVnZ.properties")
if (ac_file.Cheak() == True):
    ini = False
    if (ac_file.ReadAC() != "l2Ah5PEqxdNbaZtmeJWDi86FGCjnkYu1UIQTLsRv\n9xQYcVe56Dz0vGsLS3jWkryMIN4fEamn8X1pPhOA\nbfynRCNDuhBvXV6e7LP52pat0j134dOlKc9zoUZH\nWDSRL2G8XUripwM5VnE60zAq1Ihf7eBvl9JaZkso\n1xNiMBpQ7rjqP5wb9DLVIFoYy4su6cKJdUmOA2ET\ngepuh5i0aonMsUfJXwZz4A69NqOQTPyHGCr7ISdL\n9svLRldfiyXQrtHWTVZ078k6Ppb5FmgU14KMNDcC\nTpizKn5dAvrQFDYbymIGec8UaMtlZ6Ej10Pu3RSL\nQYGIvMqEjKg3PVt2s0p9BCT8Rx764ZnewXDzJmUd\naR9oxVrTJIU8mGls7kcPuDXF41hvEYzAW36HKeLB"):
        Ac_ban()
else:
    ini = True
#
isini_file = File("../config/isini.properties")
if (isini_file.Cheak() == False):
    if (ini == True):
        g.msgbox("这是你第一次使用广东海洋大学品行分杀手，本软件由广东海洋大学能源1214李燊同学制作",swName,"太棒啦~","../img/icon.gif")
    else:
        g.msgbox("本软件由广东海洋大学能源1214李燊同学制作", swName, "太棒啦~", "../img/icon.gif")
    g.msgbox("本软件初衷是将原本人工得花一个多小时才能解决的班级同学品行分统计工作压缩到几秒内完成，将这一任务量大且重复的过程“人工智障化“，按照软件名称顾名思义只适用于广东海洋大学\n班长学委的妈妈再也不用担心同学内卷疯狂爽，月度整理火葬场~",swName,"什么是人工智障化")
    g.msgbox("人工智障化是指将重复工作自动化的过程\n\n要想实现人工智能完全代替人工去执行非仅人类能力创新的日常性事务就需要通过深度学习自主实现人工智障化\n\n这也是我们的目标之一，期待你的加入使得我们能够在校内组建一个创业团队\n使用企业微信扫描二维码",swName,"原来如此","../img/join_us.gif")
    Ad()
    g.msgbox("接下来开始刚开始使用本软件的一些设置工作",swName,"开始吧")
    ImportClassName()
    ImportDataBasic()
    SetShowDataInclude()
    g.msgbox("一切初始化工作已完成！",swName,"开始正式使用本软件")
    isini_file.Write("1")
#
while True:
    ac_file = File("C://Users//Public//Documents//system_//DDS9QHcXbpdMjNzyOxRI1sKv3ugLmqe2JVnZ.properties")
    if (ac_file.Cheak() == False):
        workFn = g.choicebox("选择你要执行的任务。你有1次机会试用本软件，激活本软件即可永久使用",swName,("免费激活本软件","开始统计品行分","设置","反馈BUG","关于我们"))
    else:
        if (ac_file.ReadAC() == "l2Ah5PEqxdNbaZtmeJWDi86FGCjnkYu1UIQTLsRv\n9xQYcVe56Dz0vGsLS3jWkryMIN4fEamn8X1pPhOA\nbfynRCNDuhBvXV6e7LP52pat0j134dOlKc9zoUZH\nWDSRL2G8XUripwM5VnE60zAq1Ihf7eBvl9JaZkso\n1xNiMBpQ7rjqP5wb9DLVIFoYy4su6cKJdUmOA2ET\ngepuh5i0aonMsUfJXwZz4A69NqOQTPyHGCr7ISdL\n9svLRldfiyXQrtHWTVZ078k6Ppb5FmgU14KMNDcC\nTpizKn5dAvrQFDYbymIGec8UaMtlZ6Ej10Pu3RSL\nQYGIvMqEjKg3PVt2s0p9BCT8Rx764ZnewXDzJmUd\naR9oxVrTJIU8mGls7kcPuDXF41hvEYzAW36HKeLB"):
            workFn = g.choicebox("选择你要执行的任务",swName,("开始统计品行分","设置","反馈BUG","关于我们"))
        else:
            Ac_ban()
    if (workFn == '开始统计品行分'):
        while True:
            className_file = File("../config/className.properties")
            dataBasic_file = File("../config/classBasicData.xlsx")
            showDataInclude_file = File("../config/showDataInclude.properties")
            if (className_file.Cheak() == False):
                g.msgbox("班级信息配置文件丢失！", swName, "重新输入")
                ImportClassName()
            elif (dataBasic_file.Cheak() == False):
                g.msgbox("含有班级同学基本信息的Excel表格不存在！", swName, "重新导入")
                if (ImportDataBasic() == False):
                    break
            elif (showDataInclude_file.Cheak() == False):
                g.msgbox("最终生成表格统计数据包含的学生范围配置文件丢失！", swName, "重新设置")
                SetShowDataInclude()
            className_file = File("../config/className.properties")
            className = className_file.Read()
            showDataInclude_file = File("../config/showDataInclude.properties")
            showDataInclude = showDataInclude_file.Read()
            excelPath1 = "../config/classBasicData.xlsx"
            g.msgbox("接下来请导入学院品行分汇总Excel表格\n该Excel表格类似如图", swName, "好哒", "../img/help2.gif")
            while True:
                excelPath2 = g.fileopenbox("导入学院品行分汇总Excel表格", swName)
                if (excelPath2 == None):
                    break
                else:
                    try:
                        Excel(excelPath2,"Sheet1")
                    except:
                        g.msgbox("请导入Excel表格！", swName, "好的")
                    else:
                        g.msgbox("接下来请导入班级品行分汇总Excel表格模板\n该Excel表格模板类似如图", swName, "好哒", "../img/help3.gif")
                        while True:
                            excelPath3 = g.fileopenbox("导入班级品行分汇总Excel表格模板", swName)
                            if (excelPath3 == None):
                                break
                            else:
                                try:
                                    Excel(excelPath3,"Sheet1")
                                except:
                                    g.msgbox("请导入Excel表格！", swName, "好的")
                                else:
                                    g.msgbox("所有表格导入成功，接下来开始自动统计品行分汇总", swName, "好哒")
                                    #
                                    excelBasic = Excel(excelPath1, "Sheet1")
                                    num_student = excelBasic.GetRowMax() - 1
                                    dataBasic = excelBasic.GetData(2, num_student + 1, 1, 3)
                                    #
                                    excelEventClass = Excel(excelPath2, "Sheet1")
                                    num_major = (excelEventClass.GetColMax() + 1) / 5
                                    i = 0
                                    colMin = 1
                                    dataEventClass_len = 0
                                    dataEventClass = {}
                                    while i < num_major:
                                        colMax = colMin + 4
                                        dataEventClass_backup = excelEventClass.GetDataFilter(4,
                                                                                              excelEventClass.GetRowMax(),
                                                                                              colMin,
                                                                                              colMax, 0, className)
                                        for x in dataEventClass_backup:
                                            dataEventClass[(x + 1) + dataEventClass_len - 1] = dataEventClass_backup[x]
                                        dataEventClass_len += len(dataEventClass)
                                        colMin += 5
                                        i += 1
                                    excelFinal = Excel(excelPath3, "Sheet1")
                                    #
                                    rowList = 0
                                    dataBasic_delete = {}
                                    i = 0
                                    for studentList in dataBasic:
                                        studentName = dataBasic[studentList][1]
                                        num_studentEvent = 0
                                        studentEventScore = 0
                                        numAdd_studentEvent = 0
                                        for studentEventClass in dataEventClass:
                                            if (dataEventClass[studentEventClass][1] == studentName):
                                                numAdd_studentEvent += 1
                                        if (numAdd_studentEvent == 0):
                                            dataBasic[studentList][3] = ""
                                            dataBasic[studentList][4] = studentEventScore
                                            dataBasic_delete[i] = rowList
                                            i += 1
                                        else:
                                            for studentEventClass in dataEventClass:
                                                if (dataEventClass[studentEventClass][1] == studentName):
                                                    num_studentEvent += 1
                                                    if (num_studentEvent == 1):
                                                        studentEvent = EventFormat(num_studentEvent,
                                                                                   dataEventClass[studentEventClass][2],
                                                                                   dataEventClass[studentEventClass][3])
                                                    else:
                                                        studentEvent = studentEvent + "\n" + EventFormat(num_studentEvent,dataEventClass[studentEventClass][2],dataEventClass[studentEventClass][3])
                                                    studentEventScore += float(dataEventClass[studentEventClass][3])
                                            dataBasic[studentList][3] = studentEvent
                                            dataBasic[studentList][4] = studentEventScore
                                            dataBasic[rowList] = dataBasic[studentList]
                                        rowList += 1
                                    if(showDataInclude == "1"):
                                        # 删除无品行分的学生
                                        for x in dataBasic_delete:
                                            dataBasic.pop(dataBasic_delete[x])
                                        # 重新排序
                                        dataBasic_backup = {}
                                        i = 0
                                        for x in dataBasic:
                                            dataBasic_backup[i] = dataBasic[x]
                                            i += 1
                                        dataBasic = dataBasic_backup
                                        #
                                    row = 0
                                    for x in dataBasic:
                                        excelFinal.InsertNum(row + 6, 2, dataBasic[x][0])
                                        excelFinal.InsertValue(row + 6, 3, dataBasic[x][1])
                                        excelFinal.InsertValue(row + 6, 4, dataBasic[x][2])
                                        excelFinal.InsertValue(row + 6, 6, dataBasic[x][3])
                                        excelFinal.InsertValue(row + 6, 7, dataBasic[x][4])
                                        row += 1
                                    excelFinal.InsertValue(2, 3, className)
                                    excelFinal.Save(excelPath3)
                                    timeNow = datetime.datetime.now()
                                    excelPath4 = os.path.dirname(
                                        excelPath3) + "\\【" + className + "】" + str(timeNow.month - 1) + "月份品行分统计表.xlsx"
                                    os.rename(excelPath3, excelPath4)
                                    className_file.ReadClose()
                                    showDataInclude_file.ReadClose()
                                    ac_file = File(
                                        "C://Users//Public//Documents//system_//DDS9QHcXbpdMjNzyOxRI1sKv3ugLmqe2JVnZ.properties")
                                    if (ac_file.Cheak() == False):
                                        ac_file.Write(
                                            "nNiSIKoEjPx2kfuBDC9ZJgdF8XWRc0O5sz1ybth4\nnrsDtGPeMBmRFaxyXYO1LobZAEfp7igIj82lVW9w\nXeSmpo8trysHICMf5xYPLg6nbNiRVFWcjz1ZBlu9\nL6MKBIV2WNQy4kaHeGuAbxPE78rcpTZisDvJRoz9\nrPpYIZlKL2a7Owcz84MedD5hSvNsubFyq36AXRmt\nBV3e1WGsjaA9bw6IN2rX5MlLUO4nySkZfdRFDTCv\nmxu9zceIko8fZEB6ljCVTSAOgh3DwMWHU5isRJQr\nCyBYkQThe23I1qpa8vFL9bi5fxVS64NKZW0tDXdg\n2MygLehWlvFsqi04N7OPG3KafzYnZ5SVBR9rm8To\nqgvFMbtSKDGcPoT4HaY8IpshXw9mLE3irJU71Wxf")
                                    g.msgbox("品行分汇总统计完成！班级品行分汇总Excel表格已导出并覆盖原导入Excel模板", swName, "打开文件夹",
                                             "../img/icon.gif")
                                    os.system("explorer.exe " + os.path.dirname(excelPath3))
                                    Ad()
                                    break
                        break
            break
    elif (workFn == '设置'):
        while True:
            workFn_setting = g.choicebox("设置",swName,("重新输入班级","重新导入含有班级同学基本信息的Excel表格","重新设置最终生成表格统计数据包含的学生范围"))
            if (workFn_setting == '重新输入班级'):
                ImportClassName()
            elif (workFn_setting == '重新导入含有班级同学基本信息的Excel表格'):
                ImportDataBasic()
            elif(workFn_setting == '重新设置最终生成表格统计数据包含的学生范围'):
                SetShowDataInclude()
            else:
                break
    elif (workFn == '反馈BUG'):
        webbrowser.open("https://github.com/leoweyr/GDOU-Character_Score_Killer")
        webbrowser.open("https://space.bilibili.com/381580563")
        g.msgbox("BiliBili关注并私信”想学魔法的炜翼麻瓜“或Github在“Star”项目并提交”issue“",swName,"明白")
    elif (workFn == '关于我们'):
        g.msgbox("本软件初衷是将原本人工得花一个多小时才能解决的班级同学品行分统计工作压缩到几秒内完成，将这一任务量大且重复的过程“人工智障化“，按照软件名称顾名思义只适用于广东海洋大学\n班长学委的妈妈再也不用担心同学内卷疯狂爽，月度整理火葬场~\n\n人工智障化是指将重复工作自动化的过程\n\n要想实现人工智能完全代替人工去执行非仅人类能力创新的日常性事务就需要通过深度学习自主实现人工智障化\n\n这也是我们的目标之一，期待你的加入使得我们能够在校内组建一个创业团队\n使用企业微信扫描二维码",swName,"原来如此","../img/join_us.gif")
    elif (workFn == '免费激活本软件'):
        Ac()
    elif (workFn == None):
        os._exit(0)
    else:
        pass