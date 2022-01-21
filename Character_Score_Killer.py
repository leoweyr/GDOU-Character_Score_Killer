import easygui as g
import os
import openpyxl
import string
#
swName = "品行分统计杀手"
num_student = 35
#
class excel:
    def __init__(self,filePath,sheet):
        self.excelOpen = openpyxl.load_workbook(filePath)
        self.sheetOpen = self.excelOpen[sheet]
    def getData(self,min_row, max_row, min_col, max_col):
        rawLoop = 0
        sheetData = {}
        for row in self.sheetOpen.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            rowlist = {}
            i = 0
            for cell in row:
                rowlist[i] = cell.value
                i += 1
            sheetData[rawLoop] = rowlist
            rawLoop += 1
        return sheetData
    def getDataFilter(self,min_row,max_row,min_col,max_col,conditionKey,conditionValue):
        rawLoop = 0
        sheetData = {}
        for row in self.sheetOpen.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            rowlist = {}
            i = 0
            for cell in row:
                rowlist[i] = cell.value
                i += 1
            if (rowlist[conditionKey] == conditionValue):
                sheetData[rawLoop] = rowlist
                rawLoop += 1
        return sheetData
    def getRowMax(self):
        return self.sheetOpen.max_row
    def insertValue(self,row,col,v):
        self.sheetOpen.cell(row,col).value = v
    def insertNum(self,row,col,v):
        self.sheetOpen.cell(row, col).value = v
        self.sheetOpen.cell(row,col).number_format = '0'
    def save(self,filePath):
        self.excelOpen.save(filePath)

def eventFormat(id, event, score):
    global space
    count_en = count_dg = count_zh = count_pu = 0
    for word in event:
        if word in string.ascii_letters:
            count_en += 1
        elif word.isdigit():
            count_dg += 1
        elif word.isalpha():
            count_zh += 1
        else:
            count_pu += 1
    num_space = 36 - (count_en + count_dg + count_pu + (count_zh * 2))
    i = 0
    while (i < num_space):
        if (i == 0):
            space = ' '
        else:
            space = space + ' '
        i += 1
    return str(id) + '.' + event + space + str(score)
#
while True:
        c1 = g.ccbox("欢迎使用品行分统计杀手\n本软件初衷是把任务量大重复性的班级品行分统计工作“人工智障化“\n人工智障化是指将重复工作自动化的过程\n而想实现人工智能代替传统人工去执行重复性的工作需要通过深度学习自主实现人工智障化\n这也是我们的目标之一，期待你一起加入我们使得能够在校内创立一个创业团队\n©2022 广东海洋大学 - 能源1214 - 李燊.  All rights reserved. ",swName,('开始使用','加入我们'))
        if (c1 == None):
                os._exit(0)
        if (c1 == 1):
                break
        else:
                g.msgbox("使用企业微信扫描二维码加入我们",swName,"开始使用","img/join_us.gif")
                break
excelPath1 = g.fileopenbox("导入含有班级学生基本信息的Excel表格",swName)
if (excelPath1 == None):
        os._exit(0)
excelPath2 = g.fileopenbox("导入学院品行分汇总Excel表格",swName)
if (excelPath2 == None):
    os._exit(0)
excelPath3 = g.fileopenbox("导出空白的班级品行分汇总Excel表格",swName)
if (excelPath3 == None):
    os._exit(0)
excelBasic = excel(excelPath1,"Sheet1")
dataBasic = excelBasic.getData(2,num_student + 1,1,3)
excelEventClass = excel(excelPath2,"Sheet1")
dataEventClass = excelEventClass.getDataFilter(4,excelEventClass.getRowMax(),6,9,0,"能源1214")
excelFinal = excel(excelPath3,"Sheet1")
#
rowList = 0
dataBasic_delete = {}
i = 0
for studentList in dataBasic:
    studentName = dataBasic[studentList][1]
    num_studentEvent = 0
    studentEventScore = 0
    numAdd_studentEvent = 0
    for studentEventClass in dataEventClass:
        if (dataEventClass[studentEventClass][1] == studentName):
            numAdd_studentEvent += 1
    if (numAdd_studentEvent == 0):
            dataBasic_delete[i] = rowList
            i += 1
    else:
        for studentEventClass in dataEventClass:
            if (dataEventClass[studentEventClass][1] == studentName):
                num_studentEvent += 1
                if (num_studentEvent == 1):
                    studentEvent = eventFormat(num_studentEvent,dataEventClass[studentEventClass][2],dataEventClass[studentEventClass][3])
                else:
                    studentEvent = studentEvent + "\n" + eventFormat(num_studentEvent,dataEventClass[studentEventClass][2],dataEventClass[studentEventClass][3])
                studentEventScore += dataEventClass[studentEventClass][3]
        dataBasic[studentList][3] = studentEvent
        dataBasic[studentList][4] = studentEventScore
        dataBasic[rowList] = dataBasic[studentList]
    rowList += 1
#删除无品行分的学生
for x in dataBasic_delete:
    dataBasic.pop(dataBasic_delete[x])
#重新排序
dataBasic_backup = {}
i = 0
for x in dataBasic:
    dataBasic_backup[i] = dataBasic[x]
    i += 1
dataBasic = dataBasic_backup
#
row = 0
for x in dataBasic:
    excelFinal.insertNum(row + 6,2,dataBasic[x][0])
    excelFinal.insertValue(row + 6, 3, dataBasic[x][1])
    excelFinal.insertValue(row + 6, 4, dataBasic[x][2])
    excelFinal.insertValue(row + 6, 6, dataBasic[x][3])
    excelFinal.insertValue(row + 6, 7, dataBasic[x][4])
    row += 1
excelFinal.save(excelPath3)
g.msgbox("已完成任务",swName,"关闭软件")






